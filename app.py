from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///employee_management.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# 資料庫模型
class Employee(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    gender = db.Column(db.String(10))
    birth_date = db.Column(db.Date)
    marital_status = db.Column(db.String(20))
    id_number = db.Column(db.String(20))
    address = db.Column(db.Text)
    household_address = db.Column(db.Text)
    email = db.Column(db.String(120))
    contact_phone = db.Column(db.String(20))
    mobile_phone = db.Column(db.String(20))
    physical_condition = db.Column(db.Text)
    height = db.Column(db.Integer)
    weight = db.Column(db.Integer)
    special_status = db.Column(db.Text)
    disability_status = db.Column(db.Text)
    emergency_contact_name = db.Column(db.String(100))
    emergency_contact_relation = db.Column(db.String(50))
    emergency_contact_phone = db.Column(db.String(20))
    education = db.Column(db.Text)
    work_experience = db.Column(db.Text)
    family_status = db.Column(db.Text)
    specialty = db.Column(db.Text)
    hobbies = db.Column(db.Text)
    insurance = db.Column(db.Text)
    vision = db.Column(db.String(20))
    languages = db.Column(db.Text)
    driver_license = db.Column(db.Text)
    military_service = db.Column(db.String(20))
    military_branch = db.Column(db.String(50))
    certifications = db.Column(db.Text)
    financial_reports = db.Column(db.String(10))
    computer_skills = db.Column(db.Text)
    insurance_items = db.Column(db.Text)
    criminal_record = db.Column(db.Text)
    is_active = db.Column(db.Boolean, default=True)
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class AttendanceRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    clock_in_time = db.Column(db.DateTime)
    clock_out_time = db.Column(db.DateTime)
    # GPS位置資訊
    clock_in_latitude = db.Column(db.Float)
    clock_in_longitude = db.Column(db.Float)
    clock_in_address = db.Column(db.String(500))
    clock_out_latitude = db.Column(db.Float)
    clock_out_longitude = db.Column(db.Float)
    clock_out_address = db.Column(db.String(500))
    employee = db.relationship('Employee', backref='attendance_records')

@login_manager.user_loader
def load_user(user_id):
    return Employee.query.get(int(user_id))

# 路由
@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        employee = Employee.query.filter_by(username=username).first()
        
        if employee and check_password_hash(employee.password_hash, password) and employee.is_active:
            login_user(employee)
            return redirect(url_for('dashboard'))
        else:
            flash('登入失敗，請檢查帳號密碼或帳號狀態')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/employee_management')
@login_required
def employee_management():
    if not current_user.is_admin:
        flash('權限不足')
        return redirect(url_for('dashboard'))
    
    employees = Employee.query.filter_by(is_admin=False).all()
    return render_template('employee_management.html', employees=employees)

@app.route('/add_employee', methods=['GET', 'POST'])
@login_required
def add_employee():
    if not current_user.is_admin:
        flash('權限不足')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        # 調試：顯示接收到的表單數據
        print("接收到的表單數據:")
        for key, value in request.form.items():
            print(f"{key}: {value}")
        
        # 處理表單資料
        try:
            employee = Employee(
                employee_id=request.form.get('employee_id', ''),
                username=request.form.get('username', ''),
                password_hash=generate_password_hash(request.form.get('password', '')),
                name=request.form.get('name', ''),
                gender=request.form.get('gender', ''),
                birth_date=datetime.strptime(request.form.get('birth_date', ''), '%Y-%m-%d').date() if request.form.get('birth_date') else None,
                marital_status=request.form.get('marital_status', ''),
                id_number=request.form.get('id_number', ''),
                address=request.form.get('address', ''),
                household_address=request.form.get('household_address', ''),
                email=request.form.get('email', ''),
                contact_phone=request.form.get('contact_phone', ''),
                mobile_phone=request.form.get('mobile_phone', ''),
                physical_condition=request.form.get('physical_condition', ''),
                height=int(request.form.get('height', 0)) if request.form.get('height') else None,
                weight=int(request.form.get('weight', 0)) if request.form.get('weight') else None,
                special_status=request.form.get('special_status', ''),
                disability_status=request.form.get('disability_status', ''),
                emergency_contact_name=request.form.get('emergency_contact_name', ''),
                emergency_contact_relation=request.form.get('emergency_contact_relation', ''),
                emergency_contact_phone=request.form.get('emergency_contact_phone', ''),
                education=request.form.get('education', ''),
                work_experience=request.form.get('work_experience', ''),
                family_status=request.form.get('family_status', ''),
                specialty=request.form.get('specialty', ''),
                hobbies=request.form.get('hobbies', ''),
                insurance=request.form.get('insurance', ''),
                vision=request.form.get('vision', ''),
                languages=request.form.get('languages', ''),
                driver_license=request.form.get('driver_license', ''),
                military_service=request.form.get('military_service', ''),
                military_branch=request.form.get('military_branch', ''),
                certifications=request.form.get('certifications', ''),
                financial_reports=request.form.get('financial_reports', ''),
                computer_skills=request.form.get('computer_skills', ''),
                insurance_items=request.form.get('insurance_items', ''),
                criminal_record=request.form.get('criminal_record', '')
            )
            
            db.session.add(employee)
            db.session.commit()
            flash('員工新增成功')
            return redirect(url_for('employee_management'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'員工新增失敗：{str(e)}')
            return render_template('add_employee.html')
    
    return render_template('add_employee.html')

@app.route('/edit_employee/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_employee(id):
    if not current_user.is_admin:
        flash('權限不足')
        return redirect(url_for('dashboard'))
    
    employee = Employee.query.get_or_404(id)
    
    if request.method == 'POST':
        # 更新員工資料
        try:
            employee.employee_id = request.form.get('employee_id', '')
            employee.name = request.form.get('name', '')
            employee.gender = request.form.get('gender', '')
            employee.birth_date = datetime.strptime(request.form.get('birth_date', ''), '%Y-%m-%d').date() if request.form.get('birth_date') else None
            employee.marital_status = request.form.get('marital_status', '')
            employee.id_number = request.form.get('id_number', '')
            employee.address = request.form.get('address', '')
            employee.household_address = request.form.get('household_address', '')
            employee.email = request.form.get('email', '')
            employee.contact_phone = request.form.get('contact_phone', '')
            employee.mobile_phone = request.form.get('mobile_phone', '')
            employee.physical_condition = request.form.get('physical_condition', '')
            employee.height = int(request.form.get('height', 0)) if request.form.get('height') else None
            employee.weight = int(request.form.get('weight', 0)) if request.form.get('weight') else None
            employee.special_status = request.form.get('special_status', '')
            employee.disability_status = request.form.get('disability_status', '')
            employee.emergency_contact_name = request.form.get('emergency_contact_name', '')
            employee.emergency_contact_relation = request.form.get('emergency_contact_relation', '')
            employee.emergency_contact_phone = request.form.get('emergency_contact_phone', '')
            employee.education = request.form.get('education', '')
            employee.work_experience = request.form.get('work_experience', '')
            employee.family_status = request.form.get('family_status', '')
            employee.specialty = request.form.get('specialty', '')
            employee.hobbies = request.form.get('hobbies', '')
            employee.insurance = request.form.get('insurance', '')
            employee.vision = request.form.get('vision', '')
            employee.languages = request.form.get('languages', '')
            employee.driver_license = request.form.get('driver_license', '')
            employee.military_service = request.form.get('military_service', '')
            employee.military_branch = request.form.get('military_branch', '')
            employee.certifications = request.form.get('certifications', '')
            employee.financial_reports = request.form.get('financial_reports', '')
            employee.computer_skills = request.form.get('computer_skills', '')
            employee.insurance_items = request.form.get('insurance_items', '')
            employee.criminal_record = request.form.get('criminal_record', '')
            
            db.session.commit()
            flash('員工資料更新成功')
            return redirect(url_for('employee_management'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'員工資料更新失敗：{str(e)}')
            return render_template('edit_employee.html', employee=employee)
    
    return render_template('edit_employee.html', employee=employee)

@app.route('/delete_employee/<int:id>')
@login_required
def delete_employee(id):
    if not current_user.is_admin:
        flash('權限不足')
        return redirect(url_for('dashboard'))
    
    employee = Employee.query.get_or_404(id)
    db.session.delete(employee)
    db.session.commit()
    flash('員工刪除成功')
    return redirect(url_for('employee_management'))

@app.route('/batch_delete_employees', methods=['POST'])
@login_required
def batch_delete_employees():
    if not current_user.is_admin:
        flash('權限不足')
        return redirect(url_for('dashboard'))
    
    employee_ids = request.form.getlist('employee_ids')
    for employee_id in employee_ids:
        employee = Employee.query.get(employee_id)
        if employee:
            db.session.delete(employee)
    
    db.session.commit()
    flash(f'已刪除 {len(employee_ids)} 名員工')
    return redirect(url_for('employee_management'))

@app.route('/toggle_employee_status/<int:id>')
@login_required
def toggle_employee_status(id):
    if not current_user.is_admin:
        flash('權限不足')
        return redirect(url_for('dashboard'))
    
    employee = Employee.query.get_or_404(id)
    employee.is_active = not employee.is_active
    db.session.commit()
    
    status = '啟用' if employee.is_active else '停用'
    flash(f'員工 {employee.name} 已{status}')
    return redirect(url_for('employee_management'))

@app.route('/clock_in', methods=['GET', 'POST'])
@login_required
def clock_in():
    if request.method == 'POST':
        # 處理GPS打卡
        data = request.get_json()
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        address = data.get('address')
        
        # 如果沒有地址但有座標，嘗試獲取地址
        if not address and latitude and longitude:
            address = getAddressFromCoordinatesSync(latitude, longitude)
        
        today = date.today()
        existing_record = AttendanceRecord.query.filter_by(
            employee_id=current_user.id, 
            date=today
        ).first()
        
        if existing_record and existing_record.clock_in_time:
            return jsonify({'success': False, 'message': '今日已打卡上班'})
        
        if existing_record:
            existing_record.clock_in_time = datetime.now()
            existing_record.clock_in_latitude = latitude
            existing_record.clock_in_longitude = longitude
            existing_record.clock_in_address = address
        else:
            record = AttendanceRecord(
                employee_id=current_user.id,
                date=today,
                clock_in_time=datetime.now(),
                clock_in_latitude=latitude,
                clock_in_longitude=longitude,
                clock_in_address=address
            )
            db.session.add(record)
        
        db.session.commit()
        return jsonify({'success': True, 'message': '上班打卡成功'})
    
    # GET請求保持原有功能
    today = date.today()
    existing_record = AttendanceRecord.query.filter_by(
        employee_id=current_user.id, 
        date=today
    ).first()
    
    if existing_record and existing_record.clock_in_time:
        flash('今日已打卡上班')
        return redirect(url_for('dashboard'))
    
    if existing_record:
        existing_record.clock_in_time = datetime.now()
    else:
        record = AttendanceRecord(
            employee_id=current_user.id,
            date=today,
            clock_in_time=datetime.now()
        )
        db.session.add(record)
    
    db.session.commit()
    flash('上班打卡成功')
    return redirect(url_for('dashboard'))

@app.route('/clock_out', methods=['GET', 'POST'])
@login_required
def clock_out():
    if request.method == 'POST':
        # 處理GPS打卡
        data = request.get_json()
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        address = data.get('address')
        
        # 如果沒有地址但有座標，嘗試獲取地址
        if not address and latitude and longitude:
            address = getAddressFromCoordinatesSync(latitude, longitude)
        
        today = date.today()
        record = AttendanceRecord.query.filter_by(
            employee_id=current_user.id, 
            date=today
        ).first()
        
        if not record or not record.clock_in_time:
            return jsonify({'success': False, 'message': '請先打卡上班'})
        
        if record.clock_out_time:
            return jsonify({'success': False, 'message': '今日已打卡下班'})
        
        record.clock_out_time = datetime.now()
        record.clock_out_latitude = latitude
        record.clock_out_longitude = longitude
        record.clock_out_address = address
        db.session.commit()
        
        return jsonify({'success': True, 'message': '下班打卡成功'})
    
    # GET請求保持原有功能
    today = date.today()
    record = AttendanceRecord.query.filter_by(
        employee_id=current_user.id, 
        date=today
    ).first()
    
    if not record or not record.clock_in_time:
        flash('請先打卡上班')
        return redirect(url_for('dashboard'))
    
    if record.clock_out_time:
        flash('今日已打卡下班')
        return redirect(url_for('dashboard'))
    
    record.clock_out_time = datetime.now()
    db.session.commit()
    flash('下班打卡成功')
    return redirect(url_for('dashboard'))

@app.route('/attendance_records')
@login_required
def attendance_records():
    if not current_user.is_admin:
        flash('權限不足')
        return redirect(url_for('dashboard'))
    
    records = AttendanceRecord.query.join(Employee).order_by(AttendanceRecord.date.desc()).all()
    employees = Employee.query.filter_by(is_admin=False).order_by(Employee.employee_id).all()
    return render_template('attendance_records.html', records=records, employees=employees)

@app.route('/export_attendance_excel')
@login_required
def export_attendance_excel():
    if not current_user.is_admin:
        flash('權限不足')
        return redirect(url_for('dashboard'))
    
    records = AttendanceRecord.query.join(Employee).order_by(AttendanceRecord.date.desc()).all()
    return _generate_excel_file(records, "所有員工出勤記錄")

@app.route('/export_attendance_excel/<int:employee_id>')
@login_required
def export_attendance_excel_employee(employee_id):
    if not current_user.is_admin:
        flash('權限不足')
        return redirect(url_for('dashboard'))
    
    employee = Employee.query.get_or_404(employee_id)
    records = AttendanceRecord.query.filter_by(employee_id=employee_id).order_by(AttendanceRecord.date.desc()).all()
    return _generate_excel_file(records, f"{employee.name}出勤記錄")

def _generate_excel_file(records, filename_prefix):
    # 建立Excel檔案
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出勤記錄"
    
    # 標題列
    headers = ['員工編號', '員工姓名', '日期', '上班時間', '上班位置', '下班時間', '下班位置', '工作時數', '狀態']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 資料列
    for row, record in enumerate(records, 2):
        ws.cell(row=row, column=1, value=record.employee.employee_id)
        ws.cell(row=row, column=2, value=record.employee.name)
        ws.cell(row=row, column=3, value=record.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=4, value=record.clock_in_time.strftime('%H:%M:%S') if record.clock_in_time else '')
        ws.cell(row=row, column=5, value=record.clock_in_address if record.clock_in_address else '')
        ws.cell(row=row, column=6, value=record.clock_out_time.strftime('%H:%M:%S') if record.clock_out_time else '')
        ws.cell(row=row, column=7, value=record.clock_out_address if record.clock_out_address else '')
        
        # 計算工作時數
        if record.clock_in_time and record.clock_out_time:
            work_hours = (record.clock_out_time - record.clock_in_time).total_seconds() / 3600
            ws.cell(row=row, column=8, value=f"{work_hours:.2f}")
        else:
            ws.cell(row=row, column=8, value='')
        
        # 狀態
        if record.clock_in_time and record.clock_out_time:
            ws.cell(row=row, column=9, value='完整')
        elif record.clock_in_time:
            ws.cell(row=row, column=9, value='僅上班')
        else:
            ws.cell(row=row, column=9, value='未打卡')
    
    # 調整欄寬
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 儲存到記憶體
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/search_employees')
@login_required
def search_employees():
    if not current_user.is_admin:
        return jsonify({'error': '權限不足'})
    
    query = request.args.get('q', '')
    if query:
        employees = Employee.query.filter(
            Employee.name.contains(query) | 
            Employee.employee_id.contains(query) |
            Employee.username.contains(query)
        ).filter_by(is_admin=False).all()
    else:
        employees = Employee.query.filter_by(is_admin=False).all()
    
    results = []
    for emp in employees:
        results.append({
            'id': emp.id,
            'employee_id': emp.employee_id,
            'name': emp.name,
            'username': emp.username,
            'is_active': emp.is_active
        })
    
    return jsonify(results)

@app.route('/delete_attendance_record/<int:record_id>', methods=['DELETE'])
@login_required
def delete_attendance_record(record_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': '權限不足'})
    
    try:
        record = AttendanceRecord.query.get_or_404(record_id)
        db.session.delete(record)
        db.session.commit()
        return jsonify({'success': True, 'message': '出勤記錄刪除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'刪除失敗：{str(e)}'})

@app.route('/batch_delete_attendance_records', methods=['POST'])
@login_required
def batch_delete_attendance_records():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': '權限不足'})
    
    try:
        data = request.get_json()
        record_ids = data.get('record_ids', [])
        
        if not record_ids:
            return jsonify({'success': False, 'message': '未選擇要刪除的記錄'})
        
        deleted_count = 0
        for record_id in record_ids:
            record = AttendanceRecord.query.get(record_id)
            if record:
                db.session.delete(record)
                deleted_count += 1
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'成功刪除 {deleted_count} 筆出勤記錄'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'批次刪除失敗：{str(e)}'})

# 將GPS座標轉換為中文地址
async def getAddressFromCoordinates(lat, lng):
    try:
        # 使用OpenStreetMap Nominatim服務 (免費，無需API key)
        import aiohttp
        
        # 設置適當的請求頭
        headers = {
            'User-Agent': 'EmployeeManagementSystem/1.0 (https://example.com)',
            'Accept-Language': 'zh-TW,zh;q=0.9,en;q=0.8'
        }
        
        async with aiohttp.ClientSession() as session:
            url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lng}&format=json&accept-language=zh-TW"
            async with session.get(url, headers=headers) as response:
                if response.status == 200:
                    data = await response.json()
                    if 'display_name' in data:
                        # 嘗試提取更好的中文地址格式
                        if '台灣' in data['display_name'] or 'Taiwan' in data['display_name']:
                            # 解析地址組件
                            address_components = data.get('address', {})
                            
                            # 嘗試構建標準台灣地址格式
                            city = address_components.get('city', '') or address_components.get('county', '')
                            district = address_components.get('district', '') or address_components.get('suburb', '')
                            road = address_components.get('road', '')
                            house_number = address_components.get('house_number', '')
                            
                            # 構建地址字串
                            address_parts = []
                            
                            if city:
                                address_parts.append(city)
                            if district:
                                address_parts.append(district)
                            if road:
                                if house_number:
                                    address_parts.append(f"{road}{house_number}號")
                                else:
                                    address_parts.append(road)
                            
                            # 如果有完整的地址組件，使用它們
                            if address_parts:
                                return ''.join(address_parts)
                            
                            # 備用方案：從display_name提取
                            address = data['display_name']
                            parts = address.split(',')
                            taiwan_parts = []
                            for part in parts:
                                part = part.strip()
                                # 過濾掉不需要的部分
                                if any(keyword in part for keyword in ['台灣', 'Taiwan', 'Republic of China']):
                                    continue
                                if any(keyword in part for keyword in ['市', '區', '縣', '鎮', '里', '路', '街', '巷', '弄', '號']):
                                    taiwan_parts.append(part)
                            
                            if taiwan_parts:
                                # 取最相關的3-4個部分
                                return ''.join(taiwan_parts[-4:])
                        
                        # 如果不是台灣地址，嘗試簡化
                        address = data['display_name']
                        parts = address.split(',')
                        # 取最後幾個部分，避免過長的地址
                        if len(parts) > 4:
                            return ''.join(parts[-4:])
                        else:
                            return address
                    else:
                        return "地址解析失敗"
                elif response.status == 403:
                    # 如果被阻擋，使用備用服務
                    return f"座標位置 ({lat:.4f}, {lng:.4f})"
                else:
                    return "地址解析失敗"
    except Exception as e:
        # 如果主要服務失敗，使用備用格式
        return f"座標位置 ({lat:.4f}, {lng:.4f})"

# 簡化版本（同步）
def getAddressFromCoordinatesSync(lat, lng):
    try:
        import requests
        
        # 設置適當的User-Agent和請求頭
        headers = {
            'User-Agent': 'EmployeeManagementSystem/1.0 (https://example.com)',
            'Accept-Language': 'zh-TW,zh;q=0.9,en;q=0.8'
        }
        
        url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lng}&format=json&accept-language=zh-TW"
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            if 'display_name' in data:
                address = data['display_name']
                
                # 優先使用地址組件構建地址
                if 'address' in data:
                    address_components = data['address']
                    
                    # 構建標準台灣地址格式
                    address_parts = []
                    
                    # 城市/縣
                    city = address_components.get('city', '') or address_components.get('county', '')
                    if city:
                        address_parts.append(city)
                    
                    # 區/鎮
                    district = address_components.get('district', '') or address_components.get('town', '') or address_components.get('suburb', '')
                    if district:
                        address_parts.append(district)
                    
                    # 路/街
                    road = address_components.get('road', '')
                    if road:
                        house_number = address_components.get('house_number', '')
                        if house_number:
                            address_parts.append(f"{road}{house_number}")
                        else:
                            address_parts.append(road)
                    
                    # 如果有完整的地址組件，使用它們
                    if address_parts:
                        return ''.join(address_parts)
                
                # 備用方案：從display_name提取
                if '台灣' in address or 'Taiwan' in address:
                    parts = address.split(',')
                    taiwan_parts = []
                    for part in parts:
                        part = part.strip()
                        # 過濾掉不需要的部分
                        if any(keyword in part for keyword in ['台灣', 'Taiwan', 'Republic of China', '330', '10041']):
                            continue
                        if any(keyword in part for keyword in ['市', '區', '縣', '鎮', '里', '路', '街', '巷', '弄', '號']):
                            taiwan_parts.append(part)
                    
                    if taiwan_parts:
                        # 取最相關的3-4個部分，並清理多餘空格
                        clean_parts = [part.strip() for part in taiwan_parts[-4:] if part.strip()]
                        return ''.join(clean_parts)
                
                # 如果不是台灣地址，嘗試簡化
                parts = address.split(',')
                # 取最後幾個部分，避免過長的地址
                if len(parts) > 4:
                    return ''.join(parts[-4:])
                else:
                    return address
            else:
                return "地址解析失敗"
        elif response.status_code == 403:
            # 如果被阻擋，嘗試使用備用服務
            return _getAddressFromBackupService(lat, lng)
        else:
            return "地址解析失敗"
    except Exception as e:
        # 如果主要服務失敗，嘗試備用服務
        return _getAddressFromBackupService(lat, lng)

def _getAddressFromBackupService(lat, lng):
    """備用地址轉換服務"""
    try:
        # 使用Google Geocoding API的免費版本（需要API key）
        # 這裡我們使用一個簡單的地址格式
        return f"座標位置 ({lat:.4f}, {lng:.4f})"
    except:
        return f"座標位置 ({lat:.4f}, {lng:.4f})"

# 初始化資料庫和預設管理員帳號
def init_db():
    with app.app_context():
        db.create_all()
        
        # 建立預設管理員帳號
        admin = Employee.query.filter_by(username='admin').first()
        if not admin:
            admin = Employee(
                employee_id='ADMIN001',
                username='admin',
                password_hash=generate_password_hash('admin123'),
                name='系統管理員',
                is_admin=True
            )
            db.session.add(admin)
            db.session.commit()

# 初始化資料庫
init_db()

if __name__ == '__main__':
    # 本地開發時使用
    app.run(debug=True, host='0.0.0.0', port=5000)
